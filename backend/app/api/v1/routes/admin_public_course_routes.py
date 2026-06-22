"""管理员公共课程路由。"""
from __future__ import annotations

from io import BytesIO

from fastapi import APIRouter, Depends, File, Query, UploadFile
from fastapi.responses import Response
from openpyxl import Workbook, load_workbook
from sqlalchemy.orm import Session

from app.core.exceptions import BusinessException
from app.core.response import success
from app.core.security import require_role
from app.core.upload_validation import ALLOWED_EXCEL_EXTENSIONS, MAX_EXCEL_SIZE, validate_upload
from app.db.session import get_db
from app.models.entities import Question
from app.schemas.common import (
    AdminMaterialUpdate,
    AdminPublicCourseCreate,
    AdminPublicCourseUpdate,
    AdminQuestionCreate,
    AdminQuestionUpdate,
    AuthUser,
)
from app.services import admin_public_course_service as service

router = APIRouter(prefix="/public-courses", tags=["admin-public-courses"])


def _format_course(course, sync_info: dict | None = None) -> dict:
    data = {
        "id": course.id,
        "name": course.name,
        "created_at": course.created_at.isoformat() if course.created_at else "",
        "created_by": course.created_by,
        "is_public": bool(course.is_public),
        "material_count": len(course.materials),
        "question_count": len(course.questions),
    }
    if sync_info:
        data.update(sync_info)
    return data


def _format_material(material) -> dict:
    return {
        "id": material.id,
        "course_id": material.course_id,
        "course_name": material.course.name if material.course else "",
        "type": material.type,
        "title": material.title,
        "url": material.url,
        "duration": material.duration,
        "pages": material.pages,
        "size": material.size,
        "date": material.date,
        "file_id": material.file_id,
        "source_material_id": material.source_material_id,
        "is_synced": bool(material.source_material_id),
    }


def _format_question(question) -> dict:
    return {
        "id": question.id,
        "type": question.type,
        "course_id": question.course_id,
        "course_name": question.course.name if question.course else "",
        "stem": question.stem,
        "options": question.options or [],
        "answer": question.answer,
        "explanation": question.explanation or "",
        "tags": question.tags or [],
        "source_question_id": question.source_question_id,
        "is_synced": bool(question.source_question_id),
    }


@router.get("", summary="公共课程列表", description="管理员：获取所有公共课程，含同步状态摘要")
def get_public_courses(
    db: Session = Depends(get_db),
    _: AuthUser = Depends(require_role("admin")),
):
    courses = service.list_public_courses(db)
    result = []
    for course in courses:
        sync_info = service.get_course_sync_status(db, course)
        result.append(_format_course(course, sync_info))
    return success(result)


@router.post("", summary="创建公共课程", description="管理员：创建公共课程")
def add_public_course(
    data: AdminPublicCourseCreate,
    db: Session = Depends(get_db),
    current_user: AuthUser = Depends(require_role("admin")),
):
    course = service.create_public_course(db, data.name.strip(), current_user.id)
    return success(_format_course(course))


@router.put("/{course_id}", summary="编辑公共课程", description="管理员：修改公共课程名称并同步教师副本")
def edit_public_course(
    course_id: int,
    data: AdminPublicCourseUpdate,
    db: Session = Depends(get_db),
    _: AuthUser = Depends(require_role("admin")),
):
    course = service.update_public_course(db, course_id, data.name.strip())
    if not course:
        raise BusinessException(404, "公共课程不存在")
    return success(_format_course(course))


@router.delete("/{course_id}", summary="删除公共课程", description="管理员：删除公共课程")
def remove_public_course(
    course_id: int,
    db: Session = Depends(get_db),
    _: AuthUser = Depends(require_role("admin")),
):
    if not service.delete_public_course(db, course_id):
        raise BusinessException(404, "公共课程不存在")
    return success()


@router.get("/{course_id}/materials", summary="公共课程资料列表", description="管理员：获取公共课程资料")
def get_public_materials(
    course_id: int,
    db: Session = Depends(get_db),
    _: AuthUser = Depends(require_role("admin")),
):
    return success([_format_material(material) for material in service.list_public_materials(db, course_id)])


@router.post("/{course_id}/materials", summary="新增公共课程资料", description="管理员：新增资料并同步教师副本")
def add_public_material(
    course_id: int,
    data: AdminMaterialUpdate,
    db: Session = Depends(get_db),
    _: AuthUser = Depends(require_role("admin")),
):
    material = service.create_public_material(db, course_id, data.model_dump())
    return success(_format_material(material))


@router.put("/{course_id}/materials/{material_id}", summary="编辑公共课程资料", description="管理员：修改资料并同步教师副本")
def edit_public_material(
    course_id: int,
    material_id: int,
    data: AdminMaterialUpdate,
    db: Session = Depends(get_db),
    _: AuthUser = Depends(require_role("admin")),
):
    material = service.update_public_material(material_id=material_id, db=db, data=data.model_dump())
    if not material or material.course_id != course_id:
        raise BusinessException(404, "公共资料不存在")
    return success(_format_material(material))


@router.delete("/{course_id}/materials/{material_id}", summary="删除公共课程资料", description="管理员：删除资料并同步删除教师副本")
def remove_public_material(
    course_id: int,
    material_id: int,
    db: Session = Depends(get_db),
    _: AuthUser = Depends(require_role("admin")),
):
    material = service.get_public_material(db, material_id)
    if not material or material.course_id != course_id:
        raise BusinessException(404, "公共资料不存在")
    if not service.delete_public_material(db, material_id):
        raise BusinessException(404, "公共资料不存在")
    return success()


@router.get("/{course_id}/questions", summary="公共课程题库列表", description="管理员：获取公共课程题目")
def get_public_questions(
    course_id: int,
    db: Session = Depends(get_db),
    _: AuthUser = Depends(require_role("admin")),
):
    return success([_format_question(question) for question in service.list_public_questions(db, course_id)])


def _build_admin_question_template(question_type: str) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "公共题目导入模板"
    ws.append(["题型", "标签", "题干", "选项（选择题用 | 分隔）", "答案", "解析"])
    if question_type == "choice":
        ws.append(["choice", "人工智能,基础", "图灵测试由谁提出？", "A. 图灵|B. 冯·诺依曼|C. 乔布斯|D. 爱因斯坦", "A", "图灵提出了图灵测试。"])
    elif question_type == "fill":
        ws.append(["fill", "通识常识", "中国的首都是哪里？", "", "北京", "填空题直接填写答案关键词。"])
    elif question_type == "multi_choice":
        ws.append(["multi_choice", "编程基础|多选", "以下哪些是编程语言？", "A. Python|B. Java|C. HTML|D. C++", "ABD", "HTML 是标记语言，不是编程语言。"])
    else:
        ws.append(["choice", "人工智能,基础", "图灵测试由谁提出？", "A. 图灵|B. 冯·诺依曼|C. 乔布斯|D. 爱因斯坦", "A", "图灵提出了图灵测试。"])
        ws.append(["fill", "通识常识", "中国的首都是哪里？", "", "北京", "填空题直接填写答案关键词。"])
        ws.append(["multi_choice", "编程基础|多选", "以下哪些是编程语言？", "A. Python|B. Java|C. HTML|D. C++", "ABD", "HTML 是标记语言，不是编程语言。"])
    buffer = BytesIO()
    wb.save(buffer)
    return buffer.getvalue()


@router.get("/questions/import/template", summary="下载公共题目导入模板", description="管理员：下载不含课程名称列的公共题目 Excel 批量导入模板")
def download_public_question_template(
    template_type: str = Query("all", pattern="^(all|choice|fill|multi_choice)$"),
    _: AuthUser = Depends(require_role("admin")),
):
    content = _build_admin_question_template(template_type)
    filename_map = {
        "choice": "admin-choice-question-template.xlsx",
        "fill": "admin-fill-question-template.xlsx",
        "multi_choice": "admin-multi-choice-question-template.xlsx",
        "all": "admin-question-template.xlsx",
    }
    return Response(
        content=content,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f'attachment; filename="{filename_map[template_type]}"'},
    )


@router.post("/{course_id}/questions", summary="新增公共课程题目", description="管理员：新增题目并同步教师副本")
def add_public_question(
    course_id: int,
    data: AdminQuestionCreate,
    db: Session = Depends(get_db),
    _: AuthUser = Depends(require_role("admin")),
):
    question = service.create_public_question(db, course_id, data.model_dump())
    return success(_format_question(question))


@router.post("/{course_id}/questions/import", summary="Excel 批量导入公共课程题目", description="管理员：上传 Excel 批量导入题目到公共课程并同步教师副本")
def import_public_questions(
    course_id: int,
    file: UploadFile = File(...),
    db: Session = Depends(get_db),
    _: AuthUser = Depends(require_role("admin")),
):
    course = service.get_course_by_id(db, course_id)
    if not course or not course.is_public:
        raise BusinessException(404, "公共课程不存在")
    content = file.file.read()
    err = validate_upload(file.filename, len(content), allowed_extensions=ALLOWED_EXCEL_EXTENSIONS, max_size=MAX_EXCEL_SIZE)
    if err:
        raise BusinessException(400, err)
    try:
        wb = load_workbook(filename=BytesIO(content), data_only=True)
    except Exception:
        raise BusinessException(400, "Excel 文件读取失败，请确认文件是 .xlsx/.xls 格式，且没有损坏或被加密")
    ws = wb.active
    if ws.max_row < 2:
        raise BusinessException(400, "Excel 中没有可导入的题目数据，请至少保留表头并填写一行题目")
    headers = [str(c.value).strip() if c.value is not None else "" for c in next(ws.iter_rows(min_row=1, max_row=1))]
    required_header_groups = [
        ("题型", ["题型", "type"]),
        ("题干", ["题干", "stem"]),
        ("答案", ["答案", "answer"]),
    ]
    missing_headers = [
        label for label, candidates in required_header_groups if not any(candidate in headers for candidate in candidates)
    ]
    if missing_headers:
        raise BusinessException(400, f"Excel 表头缺少：{', '.join(missing_headers)}。请下载题库导入模板后按模板填写")
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if all(cell is None or str(cell).strip() == "" for cell in row):
            continue
        item = {headers[i]: row[i] if i < len(row) else None for i in range(len(headers))}
        rows.append(item)
    if not rows:
        raise BusinessException(400, "Excel 中没有可导入的题目数据，请填写题目内容后再上传")
    return success(service.import_questions_to_public_course(db, course_id, rows))


@router.put("/{course_id}/questions/{question_id}", summary="编辑公共课程题目", description="管理员：修改题目并同步教师副本")
def edit_public_question(
    course_id: int,
    question_id: int,
    data: AdminQuestionUpdate,
    db: Session = Depends(get_db),
    _: AuthUser = Depends(require_role("admin")),
):
    question = service.update_public_question(db, question_id, data.model_dump())
    if not question or question.course_id != course_id:
        raise BusinessException(404, "公共题目不存在")
    return success(_format_question(question))


@router.delete("/{course_id}/questions/{question_id}", summary="删除公共课程题目", description="管理员：删除题目并同步删除教师副本")
def remove_public_question(
    course_id: int,
    question_id: int,
    db: Session = Depends(get_db),
    _: AuthUser = Depends(require_role("admin")),
):
    question = service.get_public_question(db, question_id)
    if not question or question.course_id != course_id:
        raise BusinessException(404, "公共题目不存在")
    if not service.delete_public_question(db, question_id):
        raise BusinessException(404, "公共题目不存在")
    return success()
