"""教师端学生成绩课程化导出测试。"""
from io import BytesIO

from openpyxl import load_workbook

from app.core.security import get_password_hash
from app.models.entities import (
    Announcement,
    AnnouncementClass,
    Class,
    Course,
    Question,
    QuizAttempt,
    StudentClassEnrollment,
    TaskCompletion,
    User,
)
from tests.conftest import auth_header


def _seed_second_teacher_course(db_session):
    course = Course(name="第二课程", created_by="T001")
    db_session.add(course)
    db_session.flush()
    cls = Class(name="第二课程班", course_id=course.id, created_by="T001")
    student = User(
        id="2025003",
        name="第二学生",
        hashed_password=get_password_hash("abc123"),
        role="student",
        major="人工智能",
    )
    question = Question(
        type="choice",
        course_id=course.id,
        stem="2+2=?",
        options=["A. 3", "B. 4"],
        answer="B",
    )
    db_session.add_all([cls, student, question])
    db_session.flush()
    db_session.add(StudentClassEnrollment(user_id=student.id, class_id=cls.id, import_order=2))
    db_session.commit()
    return course, cls, student, question


def _seed_assignment(db_session, course, cls, title, question):
    ann = Announcement(course_id=course.id, teacher_id="T001", type="quiz", title=title, question_ids=[question.id])
    db_session.add(ann)
    db_session.flush()
    db_session.add(AnnouncementClass(announcement_id=ann.id, class_id=cls.id))
    db_session.commit()
    return ann


def test_task_overview_returns_course_fields_and_filters_by_course(client, db_session, teacher_token):
    first_course = db_session.query(Course).filter(Course.created_by == "T001", Course.name == "测试课程").one()
    first_class = db_session.query(Class).filter(Class.course_id == first_course.id).one()
    first_question = db_session.query(Question).filter(Question.course_id == first_course.id).one()
    second_course, second_class, _, second_question = _seed_second_teacher_course(db_session)
    _seed_assignment(db_session, first_course, first_class, "第一课程作业", first_question)
    _seed_assignment(db_session, second_course, second_class, "第二课程作业", second_question)

    all_data = client.get("/api/announcements/task-overview", headers=auth_header(teacher_token)).json()["data"]
    assert {task["course_name"] for task in all_data["tasks"]} >= {"测试课程", "第二课程"}

    filtered = client.get(
        f"/api/announcements/task-overview?course_id={second_course.id}",
        headers=auth_header(teacher_token),
    ).json()["data"]
    assert [task["title"] for task in filtered["tasks"]] == ["第二课程作业"]
    assert filtered["tasks"][0]["course_id"] == second_course.id
    assert filtered["tasks"][0]["course_name"] == "第二课程"


def test_teacher_students_filters_stats_by_course(client, db_session, teacher_token):
    first_course = db_session.query(Course).filter(Course.created_by == "T001", Course.name == "测试课程").one()
    first_class = db_session.query(Class).filter(Class.course_id == first_course.id).one()
    first_question = db_session.query(Question).filter(Question.course_id == first_course.id).one()
    second_course, second_class, second_student, second_question = _seed_second_teacher_course(db_session)
    _seed_assignment(db_session, first_course, first_class, "第一课程作业", first_question)
    second_ann = _seed_assignment(db_session, second_course, second_class, "第二课程作业", second_question)
    db_session.add(TaskCompletion(announcement_id=second_ann.id, user_id=second_student.id))
    db_session.commit()

    first_resp = client.get(
        f"/api/teacher/students?course_id={first_course.id}",
        headers=auth_header(teacher_token),
    ).json()
    assert [item["id"] for item in first_resp["data"]["items"]] == ["2025001"]
    assert first_resp["data"]["items"][0]["incomplete_tasks"] == 1

    second_resp = client.get(
        f"/api/teacher/students?course_id={second_course.id}",
        headers=auth_header(teacher_token),
    ).json()
    assert [item["id"] for item in second_resp["data"]["items"]] == ["2025003"]
    assert second_resp["data"]["items"][0]["completed_tasks"] == 1


def test_teacher_students_returns_each_task_scores(client, db_session, teacher_token):
    first_course = db_session.query(Course).filter(Course.created_by == "T001", Course.name == "测试课程").one()
    first_class = db_session.query(Class).filter(Class.course_id == first_course.id).one()
    first_question = db_session.query(Question).filter(Question.course_id == first_course.id).one()
    first_ann = _seed_assignment(db_session, first_course, first_class, "第一次作业", first_question)
    second_ann = _seed_assignment(db_session, first_course, first_class, "第二次作业", first_question)
    db_session.add_all([
        QuizAttempt(
            user_id="2025001",
            question_id=first_question.id,
            user_answer=first_question.answer,
            is_correct=True,
            announcement_id=first_ann.id,
        ),
        TaskCompletion(announcement_id=first_ann.id, user_id="2025001"),
    ])
    db_session.commit()

    resp = client.get(
        f"/api/teacher/students?course_id={first_course.id}",
        headers=auth_header(teacher_token),
    )
    assert resp.status_code == 200
    student = resp.json()["data"]["items"][0]
    assert [item["title"] for item in student["task_scores"]] == ["第一次作业", "第二次作业"]
    assert student["task_scores"][0] == {
        "announcement_id": first_ann.id,
        "title": "第一次作业",
        "score": 100,
        "is_completed": True,
    }
    assert student["task_scores"][1] == {
        "announcement_id": second_ann.id,
        "title": "第二次作业",
        "score": None,
        "is_completed": False,
    }


def test_students_export_splits_sheets_by_course_and_writes_task_scores(client, db_session, teacher_token):
    first_course = db_session.query(Course).filter(Course.created_by == "T001", Course.name == "测试课程").one()
    first_class = db_session.query(Class).filter(Class.course_id == first_course.id).one()
    first_question = db_session.query(Question).filter(Question.course_id == first_course.id).one()
    second_course, second_class, second_student, second_question = _seed_second_teacher_course(db_session)
    first_ann = _seed_assignment(db_session, first_course, first_class, "第一课程作业", first_question)
    second_ann = _seed_assignment(db_session, second_course, second_class, "第二课程作业", second_question)
    db_session.add_all([
        QuizAttempt(
            user_id="2025001",
            question_id=first_question.id,
            user_answer=first_question.answer,
            is_correct=True,
            announcement_id=first_ann.id,
        ),
        TaskCompletion(announcement_id=first_ann.id, user_id="2025001"),
        QuizAttempt(
            user_id=second_student.id,
            question_id=second_question.id,
            user_answer="A",
            is_correct=False,
            announcement_id=second_ann.id,
        ),
        TaskCompletion(announcement_id=second_ann.id, user_id=second_student.id),
    ])
    db_session.commit()

    resp = client.get("/api/teacher/students/export", headers=auth_header(teacher_token))
    assert resp.status_code == 200
    wb = load_workbook(BytesIO(resp.content))
    assert "总览" in wb.sheetnames
    assert "测试课程" in wb.sheetnames
    assert "第二课程" in wb.sheetnames

    first_sheet = wb["测试课程"]
    headers = [cell.value for cell in first_sheet[1]]
    assert "第一课程作业" in headers
    score_col = headers.index("第一课程作业") + 1
    assert first_sheet.cell(row=2, column=score_col).value == 100

    second_sheet = wb["第二课程"]
    second_headers = [cell.value for cell in second_sheet[1]]
    second_score_col = second_headers.index("第二课程作业") + 1
    assert second_sheet.cell(row=2, column=second_score_col).value == 0

    filtered_resp = client.get(
        f"/api/teacher/students/export?course_id={first_course.id}",
        headers=auth_header(teacher_token),
    )
    assert filtered_resp.status_code == 200
    filtered_wb = load_workbook(BytesIO(filtered_resp.content))
    filtered_sheet = filtered_wb[filtered_wb.sheetnames[0]]
    filtered_headers = [cell.value for cell in filtered_sheet[1]]
    assert "第一课程作业" in filtered_headers
    assert "第二课程作业" not in filtered_headers
